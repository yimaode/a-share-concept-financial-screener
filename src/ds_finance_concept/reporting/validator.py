import json
from pathlib import Path

REQUIRED_FILES = [
    "pipeline_manifest.json",
    "pdf_tables/pdf_manifest.jsonl",
    "pdf_tables/pages.jsonl",
    "evidence/evidence_hits.jsonl",
    "metrics/metric_candidates.jsonl",
    "metric_series/metric_series.jsonl",
    "metric_trends/metric_trends.jsonl",
    "concept_scores/concept_scores.json",
    "report/company_report.md",
    "report/report_manifest.json",
    "report/assets/concept_scores_bar.png",
    "report/assets/concept_scores_table.png",
    "report/assets/metric_latest_table.png",
    "final/company_analysis.xlsx",
]

REQUIRED_EXCEL_SHEETS = [
    "README", "Concept Scores", "Score Details", "Metric Candidates",
    "Metric Series", "Metric Trends", "Evidence Hits", "Review Queue", "Warnings",
]

REQUIRED_MD_SECTIONS = [
    "报告范围", "数据完整性", "概念评分总览", "财务趋势摘要",
    "证据句摘要", "风险反证", "需要人工复核", "输入文件与生成资产", "免责声明",
]

FORBIDDEN_WORDS = ["买入", "卖出", "持有", "目标价", "仓位建议"]


def validate_final_output(company_code: str, output_dir: Path) -> tuple[bool, str]:
    issues: list[str] = []
    ok: list[str] = []

    for rel_path in REQUIRED_FILES:
        p = output_dir / rel_path
        if p.exists():
            ok.append(f"  [OK] {rel_path}")
        else:
            issues.append(f"  [MISSING] {rel_path}")

    for rel_path in REQUIRED_FILES:
        p = output_dir / rel_path
        if not p.exists():
            continue
        if rel_path.endswith(".json"):
            try:
                json.loads(p.read_text(encoding="utf-8"))
                ok.append(f"  [OK] JSON valid: {rel_path}")
            except Exception as e:
                issues.append(f"  [INVALID JSON] {rel_path}: {e}")
        elif rel_path.endswith(".jsonl"):
            try:
                with p.open("r", encoding="utf-8") as f:
                    for i, line in enumerate(f, 1):
                        s = line.strip()
                        if s:
                            json.loads(s)
                ok.append(f"  [OK] JSONL valid: {rel_path}")
            except Exception as e:
                issues.append(f"  [INVALID JSONL] {rel_path} line {i}: {e}")

    xlsx_path = output_dir / "final/company_analysis.xlsx"
    if xlsx_path.exists():
        try:
            from openpyxl import load_workbook
            wb = load_workbook(xlsx_path)
            sheets = wb.sheetnames
            for s in REQUIRED_EXCEL_SHEETS:
                if s in sheets:
                    ok.append(f"  [OK] Excel sheet: {s}")
                else:
                    issues.append(f"  [MISSING] Excel sheet: {s}")
            for s in sheets:
                if s not in REQUIRED_EXCEL_SHEETS and s != "Sheet":
                    issues.append(f"  [EXTRA] Excel sheet: {s}")
        except Exception as e:
            issues.append(f"  [EXCEL ERROR] {e}")

    md_path = output_dir / "report/company_report.md"
    if md_path.exists():
        md = md_path.read_text(encoding="utf-8")
        for sec in REQUIRED_MD_SECTIONS:
            if sec in md:
                ok.append(f"  [OK] MD section: {sec}")
            else:
                issues.append(f"  [MISSING] MD section: {sec}")
        if "免责" in md or "不构成投资" in md:
            ok.append("  [OK] MD disclaimer present")
        else:
            issues.append("  [MISSING] MD disclaimer")
        for word in FORBIDDEN_WORDS:
            if word in md:
                issues.append(f"  [FORBIDDEN WORD] '{word}' found in report")

    manifest_path = output_dir / "report/report_manifest.json"
    if manifest_path.exists():
        try:
            m = json.loads(manifest_path.read_text(encoding="utf-8"))
            if "generated_files" in m:
                ok.append("  [OK] report_manifest has generated_files")
            if "skipped_assets" in m:
                ok.append("  [OK] report_manifest has skipped_assets")
        except Exception as e:
            issues.append(f"  [MANIFEST ERROR] {e}")

    pipeline_path = output_dir / "pipeline_manifest.json"
    if pipeline_path.exists():
        try:
            pm = json.loads(pipeline_path.read_text(encoding="utf-8"))
            stages = pm.get("stages", [])
            if len(stages) >= 8:
                ok.append(f"  [OK] pipeline has {len(stages)} stages")
            else:
                issues.append(f"  [PIPELINE] Only {len(stages)} stages")
            for s in stages:
                if s.get("status") == "failed" and not s.get("error_message"):
                    issues.append(f"  [PIPELINE] Failed stage '{s['stage']}' has no error_message")
        except Exception as e:
            issues.append(f"  [PIPELINE ERROR] {e}")

    passed = len(issues) == 0
    lines: list[str] = []
    lines.append(f"# 最终验收报告 — {company_code}")
    lines.append("")
    lines.append(f"## 结论: {'PASS' if passed else 'FAIL'}")
    lines.append("")
    lines.append(f"## 检查项: {len(ok)} OK, {len(issues)} issues")
    lines.append("")
    lines.append("## 通过项")
    for o in ok:
        lines.append(o)
    lines.append("")
    if issues:
        lines.append("## 问题项")
        for i in issues:
            lines.append(i)
        lines.append("")
    lines.append("## 安全边界")
    lines.append("- 无联网调用")
    lines.append("- 无 LLM 调用")
    lines.append("- 无外部 API 调用")
    lines.append("")
    lines.append("## 人工复核建议")
    lines.append("- 所有概念评分需人工确认")
    lines.append("- 财务指标候选值需人工复核")
    lines.append("- 建议用真实年报 PDF 数据重新运行完整流水线")
    lines.append("")

    return passed, "\n".join(lines)
