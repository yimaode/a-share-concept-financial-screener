import json
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter
except ImportError:
    Workbook = None


def _read_jsonl(path: Path) -> list[dict]:
    if not path.exists():
        return []
    records: list[dict] = []
    with path.open("r", encoding="utf-8") as f:
        for line in f:
            s = line.strip()
            if s:
                records.append(json.loads(s))
    return records


def _read_json(path: Path) -> list[dict]:
    if not path.exists():
        return []
    data = json.loads(path.read_text(encoding="utf-8"))
    return data.get("concepts", [])


def _write_sheet(wb, name: str, headers: list[str], rows: list[list], note: str = "") -> None:
    ws = wb.create_sheet(title=name)
    ws.append(headers)
    for row in rows:
        ws.append(row)
    ws.freeze_panes = "A2"
    if note:
        ws.insert_rows(1)
        ws["A1"] = note


def export_excel(
    company_code: str,
    metric_candidates_file: Path,
    metric_series_file: Path,
    metric_trends_file: Path,
    evidence_file: Path,
    concept_scores_file: Path,
    output_file: Path,
) -> None:
    if Workbook is None:
        raise ImportError("openpyxl is required for Excel export")

    wb = Workbook()
    ws = wb.active
    ws.title = "README"
    ws.append(["公司财务分析工作簿"])
    ws.append([f"公司代码: {company_code}"])
    ws.append(["本文件不构成投资建议。"])
    ws.append(["manual_review = 人工复核确认值. table/text = 自动抽取候选. 自动候选不等于最终财务值."])

    candidates = _read_jsonl(metric_candidates_file)
    series = _read_jsonl(metric_series_file)
    trends = _read_jsonl(metric_trends_file)
    evidence = _read_jsonl(evidence_file)
    scores = _read_json(concept_scores_file)

    ALL_SHEETS = [
        "Concept Scores", "Score Details", "Metric Candidates",
        "Metric Series", "Metric Trends", "Evidence Hits",
        "Review Queue", "Warnings",
    ]

    if candidates:
        headers = list(candidates[0].keys())
        rows = [[str(c.get(h, "")) for h in headers] for c in candidates]
    else:
        headers = ["candidate_id", "metric_id", "metric_name", "value", "report_period"]
        rows = []
    _write_sheet(wb, "Metric Candidates", headers, rows)

    if series:
        headers = list(series[0].keys())
        if "source_type" not in headers:
            headers.append("source_type")
        if "selection_method" not in headers:
            headers.append("selection_method")
        rows = [[str(s.get(h, "")) for h in headers] for s in series]
        for i, s in enumerate(series):
            if "selection_method" in s and "source_type" not in s:
                rows[i][headers.index("source_type")] = s.get("selection_method", "auto")
    else:
        headers = ["series_id", "metric_id", "metric_name", "report_period", "value_normalized", "source_type", "selection_method"]
        rows = []
    _write_sheet(wb, "Metric Series", headers, rows)

    if trends:
        headers = list(trends[0].keys())
        rows = [[str(t.get(h, "")) for h in headers] for t in trends]
    else:
        headers = ["trend_id", "metric_id", "report_period", "yoy", "cagr_3y"]
        rows = []
    _write_sheet(wb, "Metric Trends", headers, rows)

    if evidence:
        headers = list(evidence[0].keys())
        rows = [[str(e.get(h, "")) for h in headers] for e in evidence]
    else:
        headers = ["evidence_id", "concept_id", "polarity", "keyword", "sentence"]
        rows = []
    _write_sheet(wb, "Evidence Hits", headers, rows)

    if scores:
        headers = ["concept_id", "concept_name", "score", "level", "status",
                   "positive_hits", "negative_hits", "metric_coverage"]
        rows = [[str(c.get(h, "")) for h in headers] for c in scores]
    else:
        headers = ["concept_id", "concept_name", "score", "level"]
        rows = []
    _write_sheet(wb, "Concept Scores", headers, rows)

    detail_headers = ["concept_id", "component", "source_type", "points", "reason"]
    rows = [[str(c.get(h, "")) for h in detail_headers] for c in scores]
    _write_sheet(wb, "Score Details", detail_headers, rows)

    _write_sheet(wb, "Review Queue", ["item", "note"], [["人工复核", "需人工确认所有评分和候选值"]])
    _write_sheet(wb, "Warnings", ["message"], [["请检查数据完整性"]])

    output_file.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_file))
