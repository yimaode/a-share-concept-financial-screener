import json
from pathlib import Path

from ds_finance_concept.reporting.validator import validate_final_output
from ds_finance_concept.cli import main


def _w(path: Path, data, is_json=False):
    if is_json:
        path.write_text(json.dumps(data, ensure_ascii=False), encoding="utf-8")
    else:
        path.write_text(data, encoding="utf-8")


def _setup_minimal_output(tmp_path, missing_excel=False):
    out = tmp_path / "out"
    (out / "pdf_tables").mkdir(parents=True)
    (out / "evidence").mkdir(parents=True)
    (out / "metrics").mkdir(parents=True)
    (out / "metric_series").mkdir(parents=True)
    (out / "metric_trends").mkdir(parents=True)
    (out / "concept_scores").mkdir(parents=True)
    (out / "report" / "assets").mkdir(parents=True)
    (out / "final").mkdir(parents=True)

    _w(out / "pipeline_manifest.json", {"stages": [
        {"stage": f"s{i}", "status": "success", "error_message": ""} for i in range(8)
    ] + [{"stage": "s9", "status": "success", "error_message": ""}]}, True)
    _w(out / "pdf_tables/pdf_manifest.jsonl", "")
    _w(out / "pdf_tables/pages.jsonl", "")
    _w(out / "evidence/evidence_hits.jsonl", "")
    _w(out / "metrics/metric_candidates.jsonl", "")
    _w(out / "metric_series/metric_series.jsonl", "")
    _w(out / "metric_trends/metric_trends.jsonl", "")
    _w(out / "concept_scores/concept_scores.json", {"concepts": []}, True)
    _w(out / "report/company_report.md",
        "# 报告\n## 1. 报告范围\n## 2. 数据完整性\n## 3. 概念评分总览\n"
        "## 4. 财务趋势摘要\n## 5. 证据句摘要\n## 6. 风险反证\n"
        "## 7. 需要人工复核\n## 8. 输入文件与生成资产\n## 9. 免责声明\n"
        "本报告不构成投资建议。")
    _w(out / "report/report_manifest.json", {"generated_files": [], "skipped_assets": []}, True)
    _w(out / "report/assets/concept_scores_bar.png", "")
    _w(out / "report/assets/concept_scores_table.png", "")
    _w(out / "report/assets/metric_latest_table.png", "")

    if not missing_excel:
        try:
            from openpyxl import Workbook
            wb = Workbook()
            for s in ["README", "Concept Scores", "Score Details", "Metric Candidates",
                      "Metric Series", "Metric Trends", "Evidence Hits", "Review Queue", "Warnings"]:
                ws = wb.create_sheet(title=s)
                ws.append(["header"])
            wb.save(str(out / "final/company_analysis.xlsx"))
        except Exception:
            pass

    return out


def test_validator_all_pass(tmp_path):
    out = _setup_minimal_output(tmp_path)
    passed, report = validate_final_output("000001", out)
    assert passed


def test_validator_missing_file_fails(tmp_path):
    out = _setup_minimal_output(tmp_path)
    (out / "report/assets/metric_latest_table.png").unlink()
    passed, report = validate_final_output("000001", out)
    assert not passed
    assert "MISSING" in report


def test_validator_missing_excel_sheet_fails(tmp_path):
    out = _setup_minimal_output(tmp_path, missing_excel=True)
    _w(out / "final/company_analysis.xlsx", "")
    passed, report = validate_final_output("000001", out)
    assert not passed


def test_validator_forbidden_word_fails(tmp_path):
    out = _setup_minimal_output(tmp_path)
    _w(out / "report/company_report.md", "建议买入该股票。", )
    passed, report = validate_final_output("000001", out)
    assert not passed
    assert "FORBIDDEN" in report


def test_validator_cli_pass(tmp_path):
    out = _setup_minimal_output(tmp_path)
    report = tmp_path / "report.md"
    ret = main(["validate-final-output", "--company-code", "000001",
                "--output-dir", str(out), "--report-file", str(report)])
    assert ret == 0


def test_validator_cli_fail(tmp_path):
    out = _setup_minimal_output(tmp_path)
    (out / "report/assets/concept_scores_bar.png").unlink()
    report = tmp_path / "report.md"
    ret = main(["validate-final-output", "--company-code", "000001",
                "--output-dir", str(out), "--report-file", str(report)])
    assert ret == 1


def test_excel_nine_sheets(tmp_path):
    from ds_finance_concept.reporting.excel_exporter import export_excel
    for f in ["c.jsonl", "s.jsonl", "t.jsonl", "e.jsonl"]:
        (tmp_path / f).write_text("", encoding="utf-8")
    _w(tmp_path / "scores.json", {"concepts": []}, True)
    out = tmp_path / "out.xlsx"
    export_excel("000001", tmp_path / "c.jsonl", tmp_path / "s.jsonl",
                 tmp_path / "t.jsonl", tmp_path / "e.jsonl", tmp_path / "scores.json", out)

    from openpyxl import load_workbook
    wb = load_workbook(out)
    assert "README" in wb.sheetnames
    assert "Metric Candidates" in wb.sheetnames
    assert "Metric Series" in wb.sheetnames
    assert "Metric Trends" in wb.sheetnames
    assert "Evidence Hits" in wb.sheetnames
    assert "Concept Scores" in wb.sheetnames
    assert "Score Details" in wb.sheetnames
    assert "Review Queue" in wb.sheetnames
    assert "Warnings" in wb.sheetnames


def test_report_metric_latest_table(tmp_path):
    from ds_finance_concept.reporting.report_builder import build_company_report
    concepts = tmp_path / "c.json"
    _w(concepts, {"version": "0.1.0", "status": "frozen", "concepts": []}, True)
    for f in ["t.jsonl", "s.jsonl", "e.jsonl"]:
        (tmp_path / f).write_text("", encoding="utf-8")
    _w(tmp_path / "scores.json", {"concepts": []}, True)
    _, _, _ = build_company_report("000001", concepts,
        tmp_path / "t.jsonl", tmp_path / "s.jsonl", tmp_path / "e.jsonl",
        tmp_path / "scores.json", tmp_path / "report")
    assert (tmp_path / "report/assets/metric_latest_table.png").exists()


def test_report_manifest_skipped_assets(tmp_path):
    from ds_finance_concept.reporting.report_builder import build_company_report
    concepts = tmp_path / "c.json"
    _w(concepts, {"version": "0.1.0", "status": "frozen", "concepts": []}, True)
    for f in ["t.jsonl", "s.jsonl", "e.jsonl"]:
        (tmp_path / f).write_text("", encoding="utf-8")
    _w(tmp_path / "scores.json", {"concepts": []}, True)
    _, _, _ = build_company_report("000001", concepts,
        tmp_path / "t.jsonl", tmp_path / "s.jsonl", tmp_path / "e.jsonl",
        tmp_path / "scores.json", tmp_path / "report")
    m = json.loads((tmp_path / "report/report_manifest.json").read_text(encoding="utf-8"))
    assert "skipped_assets" in m


def test_pipeline_manifest_8_stages(tmp_path):
    out = _setup_minimal_output(tmp_path)
    pm = json.loads((out / "pipeline_manifest.json").read_text(encoding="utf-8"))
    assert len(pm["stages"]) >= 8
