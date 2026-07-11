import json
from pathlib import Path

import pytest

from ds_finance_concept.reporting.report_builder import build_company_report
from ds_finance_concept.reporting.report_builder import _configure_chart_font
from ds_finance_concept.reporting.excel_exporter import export_excel


def _w_json(path: Path, data) -> None:
    path.write_text(json.dumps(data, ensure_ascii=False), encoding="utf-8")


def _w_jsonl(path: Path, data: list[dict]) -> None:
    lines = [json.dumps(d, ensure_ascii=False) for d in data]
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def test_build_report_empty(tmp_path):
    concepts = tmp_path / "c.json"
    _w_json(concepts, {"version": "0.1.0", "status": "frozen", "concepts": []})
    for f in ["trends.jsonl", "series.jsonl", "evidence.jsonl", "scores.json"]:
        (tmp_path / f).write_text("" if f.endswith("jsonl") else '{"concepts":[]}', encoding="utf-8")

    md, files, _ = build_company_report(
        "000001", concepts,
        tmp_path / "trends.jsonl", tmp_path / "series.jsonl",
        tmp_path / "evidence.jsonl", tmp_path / "scores.json",
        tmp_path / "report",
    )
    assert "免责声明" in md
    assert "本报告不构成投资建议" in md
    assert (tmp_path / "report" / "company_report.md").exists()
    assert (tmp_path / "report" / "report_manifest.json").exists()


def test_chart_font_configuration_disables_unicode_minus_fallback():
    import matplotlib

    _configure_chart_font()
    assert matplotlib.rcParams["axes.unicode_minus"] is False


def test_report_no_buy_sell_words(tmp_path):
    concepts = tmp_path / "c.json"
    _w_json(concepts, {"version": "0.1.0", "status": "frozen", "concepts": []})
    for f in ["trends.jsonl", "series.jsonl", "evidence.jsonl", "scores.json"]:
        (tmp_path / f).write_text("" if f.endswith("jsonl") else '{"concepts":[]}', encoding="utf-8")

    md, _, _ = build_company_report("000001", concepts,
        tmp_path / "trends.jsonl", tmp_path / "series.jsonl",
        tmp_path / "evidence.jsonl", tmp_path / "scores.json",
        tmp_path / "report")
    for word in ["买入", "卖出", "持有", "目标价", "仓位建议"]:
        assert word not in md


def test_build_report_with_data(tmp_path):
    concepts = tmp_path / "c.json"
    _w_json(concepts, {
        "version": "0.1.0", "status": "frozen",
        "concepts": [{"concept_id": "sg", "name": "超级成长股", "hard_metrics": [], "scoring": {}}]
    })
    _w_jsonl(tmp_path / "trends.jsonl", [
        {"metric_id": "revenue", "metric_name": "营业收入", "report_period": "2024Q1",
         "period_year": 2024, "period_type": "Q1", "period_order": 1,
         "value_normalized": 100, "value_unit_normalized": "CNY", "is_percent": False,
         "yoy": 0.2, "yoy_status": "computed", "consecutive_growth_count": 1,
         "source_candidate_id": "mc_1", "source_pdf": "r.pdf", "page_number": 1},
        {"metric_id": "revenue", "metric_name": "营业收入", "report_period": "2024H1",
         "period_year": 2024, "period_type": "H1", "period_order": 2,
         "value_normalized": 200, "value_unit_normalized": "CNY", "is_percent": False,
         "yoy": 0.25, "yoy_status": "computed", "consecutive_growth_count": 2,
         "source_candidate_id": "mc_2", "source_pdf": "r.pdf", "page_number": 2},
    ])
    _w_jsonl(tmp_path / "series.jsonl", [])
    _w_jsonl(tmp_path / "evidence.jsonl", [
        {"evidence_id": "ev_1", "concept_id": "sg", "polarity": "positive",
         "keyword": "成长", "sentence": "公司持续成长", "negation_detected": False}
    ])
    _w_json(tmp_path / "scores.json", {
        "concepts": [{"concept_id": "sg", "concept_name": "超级成长股",
                      "score": 75, "level": "medium", "positive_hits": 1, "negative_hits": 0}]
    })

    md, files, _ = build_company_report("000001", concepts,
        tmp_path / "trends.jsonl", tmp_path / "series.jsonl",
        tmp_path / "evidence.jsonl", tmp_path / "scores.json",
        tmp_path / "report")
    assert "公司持续成长" in md
    assert len(files) > 0

    assets = list((tmp_path / "report" / "assets").glob("*.png"))
    assert len(assets) > 0


def test_report_manifest(tmp_path):
    concepts = tmp_path / "c.json"
    _w_json(concepts, {"version": "0.1.0", "status": "frozen", "concepts": []})
    for f in ["trends.jsonl", "series.jsonl", "evidence.jsonl", "scores.json"]:
        (tmp_path / f).write_text("" if f.endswith("jsonl") else '{"concepts":[]}', encoding="utf-8")

    _, _, _ = build_company_report("000001", concepts,
        tmp_path / "trends.jsonl", tmp_path / "series.jsonl",
        tmp_path / "evidence.jsonl", tmp_path / "scores.json",
        tmp_path / "report")
    m = json.loads((tmp_path / "report" / "report_manifest.json").read_text(encoding="utf-8"))
    assert "generated_files" in m
    assert "warnings" in m


def test_export_excel_with_data(tmp_path):
    for f, data in [
        ("candidates.jsonl", [{"candidate_id": "mc_1", "metric_id": "revenue", "value": 100}]),
        ("series.jsonl", [{"series_id": "ms_1", "metric_id": "revenue", "value_normalized": 100}]),
        ("trends.jsonl", [{"trend_id": "mt_1", "metric_id": "revenue"}]),
        ("evidence.jsonl", [{"evidence_id": "ev_1", "concept_id": "sg"}]),
        ("scores.json", {"concepts": [{"concept_id": "sg", "score": 50}]}),
    ]:
        path = tmp_path / f
        if f.endswith("jsonl"):
            _w_jsonl(path, data)
        else:
            _w_json(path, data)

    out = tmp_path / "out.xlsx"
    export_excel("000001",
        tmp_path / "candidates.jsonl", tmp_path / "series.jsonl",
        tmp_path / "trends.jsonl", tmp_path / "evidence.jsonl",
        tmp_path / "scores.json", out)
    assert out.exists()

    from openpyxl import load_workbook
    wb = load_workbook(out)
    sheets = wb.sheetnames
    assert "README" in sheets
    assert "Concept Scores" in sheets
    assert "Metric Candidates" in sheets
    assert "Evidence Hits" in sheets


def test_export_excel_empty(tmp_path):
    for f in ["candidates.jsonl", "series.jsonl", "trends.jsonl", "evidence.jsonl"]:
        (tmp_path / f).write_text("", encoding="utf-8")
    _w_json(tmp_path / "scores.json", {"concepts": []})

    out = tmp_path / "out.xlsx"
    export_excel("000001",
        tmp_path / "candidates.jsonl", tmp_path / "series.jsonl",
        tmp_path / "trends.jsonl", tmp_path / "evidence.jsonl",
        tmp_path / "scores.json", out)
    assert out.exists()

    from openpyxl import load_workbook
    wb = load_workbook(out)
    assert "README" in wb.sheetnames


def test_excel_disclaimer(tmp_path):
    out = tmp_path / "out.xlsx"
    for f in ["candidates.jsonl", "series.jsonl", "trends.jsonl", "evidence.jsonl"]:
        (tmp_path / f).write_text("", encoding="utf-8")
    _w_json(tmp_path / "scores.json", {"concepts": []})
    export_excel("000001",
        tmp_path / "candidates.jsonl", tmp_path / "series.jsonl",
        tmp_path / "trends.jsonl", tmp_path / "evidence.jsonl",
        tmp_path / "scores.json", out)

    from openpyxl import load_workbook
    wb = load_workbook(out)
    assert "不构成投资建议" in wb["README"]["A3"].value
